import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from config.settings import Settings
from config.constants import AttendanceStatus


class AttendanceRecorder:
    def __init__(self):
        os.makedirs(Settings.ATTENDANCE_RECORDS_DIR, exist_ok=True)
        self.month_dir = os.path.join(Settings.ATTENDANCE_RECORDS_DIR,
                                      datetime.now().strftime("%Y-%m"))
        os.makedirs(self.month_dir, exist_ok=True)
        self.filename = os.path.join(self.month_dir,
                                     f"{datetime.now().strftime('%Y-%m-%d')}.xlsx")

    def record(self, all_students, present_students):
        """记录考勤结果到Excel"""
        if os.path.exists(self.filename):
            wb = load_workbook(self.filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["姓名", "状态", "时间"])

        for student in all_students:
            status = AttendanceStatus.PRESENT if student in present_students else AttendanceStatus.ABSENT
            ws.append([student, status, datetime.now().strftime("%H:%M:%S")])

        wb.save(self.filename)